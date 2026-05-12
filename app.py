import os, uuid, csv, json, base64, io
from datetime import datetime, date, timedelta
from flask import Flask, render_template, redirect, url_for, flash, request, session, jsonify, make_response
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import pyotp
import qrcode
from config import Config
from models import db, User, Role, Branch, Product, Category, Supplier, Purchase, Sale, SaleItem, StockTransfer, ExpiryAlert
from forms import LoginForm, ProductForm, SupplierForm, PurchaseForm, StockTransferForm, UserForm, CategoryForm
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

app = Flask(__name__)
app.config.from_object(Config)
db.init_app(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# ---------- Role helpers ----------
def roles_required(*roles):
    def decorator(f):
        @login_required
        def wrapped(*args, **kwargs):
            if current_user.role.name not in roles:
                flash('Access denied.')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        wrapped.__name__ = f.__name__
        return wrapped
    return decorator

def get_active_branch():
    bid = session.get('active_branch_id')
    if bid:
        return Branch.query.get(bid)
    if not current_user.is_anonymous and not current_user.is_admin():
        return current_user.branch
    return None

# ---------- Auth ----------
@app.route('/')
def index():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET','POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user and check_password_hash(user.password_hash, form.password.data) and user.is_active:
            # ---- NEW: Force admin 2FA setup if not yet enrolled ----
            if user.is_admin() and not user.totp_secret:
                # Generate a temporary secret (not saved yet)
                session['pending_totp_secret'] = pyotp.random_base32()
                session['pending_user_id'] = user.id
                # Show the setup page immediately – user is not logged in yet
                return redirect(url_for('first_time_2fa'))

            # Normal login: check 2FA if already enrolled
            if user.totp_secret:
                if not form.otp_token.data:
                    flash('Please enter your 2FA code.')
                    return render_template('login.html', form=form)
                totp = pyotp.TOTP(user.totp_secret)
                if not totp.verify(form.otp_token.data):
                    flash('Invalid 2FA code.')
                    return render_template('login.html', form=form)

            login_user(user)
            flash('Welcome, ' + user.username)
            return redirect(url_for('dashboard'))
        flash('Invalid credentials or account disabled.')
    return render_template('login.html', form=form)

# ---- NEW ROUTE for admin's first 2FA setup ----
@app.route('/first_time_2fa', methods=['GET', 'POST'])
def first_time_2fa():
    # Must have a pending setup
    if 'pending_totp_secret' not in session or 'pending_user_id' not in session:
        flash('No pending 2FA setup. Please login again.')
        return redirect(url_for('login'))

    temp_secret = session['pending_totp_secret']
    user = User.query.get(session['pending_user_id'])
    if not user:
        # Should not happen, but clear and redirect
        session.pop('pending_totp_secret', None)
        session.pop('pending_user_id', None)
        flash('User not found.')
        return redirect(url_for('login'))

    if request.method == 'POST':
        otp = request.form.get('otp_token')
        totp = pyotp.TOTP(temp_secret)
        if totp.verify(otp):
            # OTP correct: save the secret to database
            user.totp_secret = temp_secret
            db.session.commit()
            # Clear pending data
            session.pop('pending_totp_secret', None)
            session.pop('pending_user_id', None)
            # Log the admin in
            login_user(user)
            flash('2FA setup complete. You are now logged in.')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid code. Please try again.')

    # GET: show QR code and OTP input
    totp = pyotp.TOTP(temp_secret)
    provisioning_uri = totp.provisioning_uri(name=user.username, issuer_name='SAMS Clinic POS')
    img = qrcode.make(provisioning_uri)
    buffer = io.BytesIO()
    img.save(buffer, format='PNG')
    img_str = base64.b64encode(buffer.getvalue()).decode()
    return render_template('first_time_2fa.html', qr_code=img_str, secret=temp_secret)

@app.route('/logout')
def logout():
    logout_user()
    session.pop('active_branch_id', None)
    session.pop('pending_totp_secret', None)
    session.pop('pending_user_id', None)
    return redirect(url_for('login'))

# ---------- Dashboard (role-based redirect) ----------
@app.route('/dashboard')
@login_required
def dashboard():
    if current_user.is_admin():
        return redirect(url_for('admin_dashboard'))
    elif current_user.is_cashier():
        return redirect(url_for('cashier_dashboard'))
    elif current_user.is_pharmacist():
        return redirect(url_for('pharmacist_dashboard'))
    flash('Role not defined.')
    return redirect(url_for('login'))

# ---------- Admin Dashboard ----------
@app.route('/admin_dashboard')
@login_required
@roles_required('admin')
def admin_dashboard():
    branch = get_active_branch()
    if not branch:
        flash('Select a branch first.')
        return redirect(url_for('select_branch'))
    today = date.today()
    today_sales = Sale.query.filter(Sale.branch_id==branch.id, db.func.date(Sale.sale_date)==today).count()
    monthly_sales = Sale.query.filter(Sale.branch_id==branch.id, db.extract('month', Sale.sale_date)==today.month).count()
    total_revenue_today = db.session.query(db.func.sum(Sale.total_amount)).filter(Sale.branch_id==branch.id, db.func.date(Sale.sale_date)==today).scalar() or 0
    today_profit = db.session.query(db.func.sum(SaleItem.profit)).join(Sale).filter(
        Sale.branch_id==branch.id, 
        db.func.date(Sale.sale_date)==today
    ).scalar() or 0
    monthly_profit = db.session.query(db.func.sum(SaleItem.profit)).join(Sale).filter(
        Sale.branch_id==branch.id,
        db.extract('month', Sale.sale_date)==today.month,
        db.extract('year', Sale.sale_date)==today.year
    ).scalar() or 0
    top_product = db.session.query(
        Product.name,
        db.func.sum(SaleItem.profit).label('total_profit')
    ).join(SaleItem).join(Sale).filter(
        Sale.branch_id==branch.id,
        db.extract('month', Sale.sale_date)==today.month,
        db.extract('year', Sale.sale_date)==today.year
    ).group_by(Product.id, Product.name).order_by(db.desc('total_profit')).first()
    low_stock = Product.query.filter(Product.branch_id==branch.id, Product.quantity < 10).count()
    expiring = Product.query.filter(Product.branch_id==branch.id, Product.expiry_date <= today + timedelta(days=90), Product.expiry_date >= today).count()
    total_products = Product.query.filter_by(branch_id=branch.id).count()
    active_staff = User.query.filter(User.is_active==True, User.role.has(name='cashier')).count()
    return render_template('admin_dashboard.html', branch=branch, today_sales=today_sales,
                           monthly_sales=monthly_sales, total_revenue_today=total_revenue_today,
                           today_profit=today_profit, monthly_profit=monthly_profit, 
                           top_product=top_product, low_stock=low_stock, expiring=expiring, 
                           total_products=total_products, active_staff=active_staff)

# ---------- Cashier Dashboard ----------
@app.route('/cashier_dashboard')
@login_required
@roles_required('cashier')
def cashier_dashboard():
    branch = get_active_branch()
    if not branch:
        return redirect(url_for('select_branch'))
    today = date.today()
    recent_sales = Sale.query.filter(Sale.user_id==current_user.id, db.func.date(Sale.sale_date)==today).order_by(Sale.sale_date.desc()).limit(5).all()
    return render_template('cashier_dashboard.html', branch=branch, recent_sales=recent_sales)

# ---------- Pharmacist Dashboard ----------
@app.route('/pharmacist_dashboard')
@login_required
@roles_required('pharmacist')
def pharmacist_dashboard():
    branch = get_active_branch()
    if not branch:
        return redirect(url_for('select_branch'))
    low_stock = Product.query.filter(Product.branch_id==branch.id, Product.quantity < 10).count()
    expiring = Product.query.filter(Product.branch_id==branch.id, Product.expiry_date <= date.today() + timedelta(days=90), Product.expiry_date >= date.today()).count()
    return render_template('pharmacist_dashboard.html', branch=branch, low_stock=low_stock, expiring=expiring)

# ---------- Branch selection (admin only) ----------
@app.route('/select_branch', methods=['GET','POST'])
@login_required
def select_branch():
    if not current_user.is_admin():
        if current_user.branch:
            session['active_branch_id'] = current_user.branch.id
            return redirect(url_for('dashboard'))
        else:
            flash('No branch assigned. Contact admin.')
            return redirect(url_for('logout'))
    branches = Branch.query.all()
    if request.method == 'POST':
        bid = request.form.get('branch_id')
        if bid:
            session['active_branch_id'] = int(bid)
            return redirect(url_for('dashboard'))
    return render_template('select_branch.html', branches=branches)

# ---------- Products CRUD ----------
@app.route('/products')
@login_required
@roles_required('admin','pharmacist')
def products():
    branch = get_active_branch()
    if not branch:
        return redirect(url_for('select_branch'))
    products = Product.query.filter_by(branch_id=branch.id).order_by(Product.name).all()
    return render_template('products.html', branch=branch, products=products, now=date.today())

@app.route('/add_product', methods=['GET','POST'])
@login_required
@roles_required('admin','pharmacist')
def add_product():
    branch = get_active_branch()
    if not branch:
        return redirect(url_for('select_branch'))
    form = ProductForm()
    form.category_id.choices = [(c.id, c.name) for c in Category.query.all()]
    form.supplier_id.choices = [(0, 'None')] + [(s.id, s.name) for s in Supplier.query.all()]
    if form.validate_on_submit():
        product = Product(
            branch_id=branch.id,
            name=form.name.data,
            barcode=form.barcode.data or None,
            batch_number=form.batch_number.data or None,
            category_id=form.category_id.data if form.category_id.data else None,
            supplier_id=form.supplier_id.data if form.supplier_id.data != 0 else None,
            purchase_price=form.purchase_price.data,
            selling_price=form.selling_price.data,
            quantity=form.quantity.data,
            expiry_date=form.expiry_date.data
        )
        db.session.add(product)
        db.session.commit()
        flash('Product added.')
        return redirect(url_for('products'))
    return render_template('add_product.html', form=form, branch=branch)

@app.route('/edit_product/<int:id>', methods=['GET','POST'])
@login_required
@roles_required('admin','pharmacist')
def edit_product(id):
    product = Product.query.get_or_404(id)
    form = ProductForm(obj=product)
    form.category_id.choices = [(c.id, c.name) for c in Category.query.all()]
    form.supplier_id.choices = [(0, 'None')] + [(s.id, s.name) for s in Supplier.query.all()]
    if form.validate_on_submit():
        product.name = form.name.data
        product.barcode = form.barcode.data
        product.batch_number = form.batch_number.data
        product.category_id = form.category_id.data if form.category_id.data else None
        product.supplier_id = form.supplier_id.data if form.supplier_id.data != 0 else None
        product.purchase_price = form.purchase_price.data
        product.selling_price = form.selling_price.data
        product.quantity = form.quantity.data
        product.expiry_date = form.expiry_date.data
        db.session.commit()
        flash('Product updated.')
        return redirect(url_for('products'))
    return render_template('edit_product.html', form=form, branch=get_active_branch(), product=product)

@app.route('/delete_product/<int:product_id>', methods=['DELETE'])
@login_required
@roles_required('admin','pharmacist')
def delete_product(product_id):
    branch = get_active_branch()
    if not branch:
        return jsonify({'error': 'Branch not set'}), 400
    product = Product.query.filter_by(id=product_id, branch_id=branch.id).first()
    if not product:
        return jsonify({'error': 'Product not found'}), 404
    sales_count = SaleItem.query.filter_by(product_id=product_id).count()
    if sales_count > 0:
        return jsonify({'error': 'Cannot delete product with sales history.'}), 400
    try:
        db.session.delete(product)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Failed to delete product'}), 500

# ---------- POS (Cashier) ----------
@app.route('/pos')
@login_required
@roles_required('cashier')
def pos():
    branch = get_active_branch()
    if not branch:
        return redirect(url_for('select_branch'))
    return render_template('pos.html', branch=branch)

@app.route('/api/search_product')
@login_required
@roles_required('cashier')
def search_product():
    query = request.args.get('q', '').strip()
    branch = get_active_branch()
    if not branch:
        return jsonify([])
    products = Product.query.filter(
        Product.branch_id==branch.id,
        Product.quantity > 0,
        (Product.name.ilike(f'%{query}%') | (Product.barcode == query) if query else True)
    ).limit(20).all()
    results = [{'id':p.id, 'name':p.name, 'barcode':p.barcode, 'selling_price':float(p.selling_price), 'stock':p.quantity} for p in products]
    return jsonify(results)

@app.route('/api/checkout', methods=['POST'])
@login_required
@roles_required('cashier')
def checkout():
    branch = get_active_branch()
    if not branch:
        return jsonify({'error':'Branch not set'}), 400
    data = request.get_json()
    if not data or 'items' not in data or not data['items']:
        return jsonify({'error':'No items'}), 400
    payment_method = data.get('payment_method', 'cash')
    total_amount = data.get('total_amount', 0)
    sale = Sale(
        branch_id=branch.id,
        user_id=current_user.id,
        total_amount=total_amount,
        payment_method=payment_method,
        cash_amount=data.get('cash_amount', 0.0),
        mpesa_amount=data.get('mpesa_amount', 0.0),
        amount_received=data.get('amount_received', total_amount),
        balance_given=data.get('balance_given', 0.0),
        mpesa_code=data.get('mpesa_code'),
        receipt_number=uuid.uuid4().hex[:10].upper()
    )
    db.session.add(sale)
    for item in data['items']:
        product = Product.query.get(item['id'])
        if not product or product.branch_id != branch.id:
            return jsonify({'error':f'Product {item["id"]} not found'}), 400
        qty = int(item['qty'])
        if qty > product.quantity:
            return jsonify({'error':f'Insufficient stock for {product.name} (only {product.quantity} left)'}), 400
        price = product.selling_price
        profit = (price - product.purchase_price) * qty
        sale_item = SaleItem(
            sale_id=sale.id,
            product_id=product.id,
            quantity=qty,
            price=price,
            profit=profit
        )
        product.quantity -= qty
        db.session.add(sale_item)
    db.session.commit()
    return jsonify({'receipt_id':sale.id, 'receipt_number':sale.receipt_number})

@app.route('/receipt/<int:sale_id>')
@login_required
def receipt(sale_id):
    sale = Sale.query.get_or_404(sale_id)
    return render_template('receipt.html', sale=sale)

# ---------- Suppliers CRUD ----------
@app.route('/suppliers')
@login_required
@roles_required('admin','pharmacist')
def suppliers():
    suppliers = Supplier.query.all()
    return render_template('suppliers.html', suppliers=suppliers)

@app.route('/add_supplier', methods=['GET','POST'])
@login_required
@roles_required('admin','pharmacist')
def add_supplier():
    form = SupplierForm()
    if form.validate_on_submit():
        supplier = Supplier(name=form.name.data, phone=form.phone.data, email=form.email.data, address=form.address.data)
        db.session.add(supplier)
        db.session.commit()
        flash('Supplier added.')
        return redirect(url_for('suppliers'))
    return render_template('add_supplier.html', form=form)

@app.route('/edit_supplier/<int:id>', methods=['GET','POST'])
@login_required
@roles_required('admin','pharmacist')
def edit_supplier(id):
    supplier = Supplier.query.get_or_404(id)
    form = SupplierForm(obj=supplier)
    if form.validate_on_submit():
        supplier.name = form.name.data
        supplier.phone = form.phone.data
        supplier.email = form.email.data
        supplier.address = form.address.data
        db.session.commit()
        flash('Supplier updated.')
        return redirect(url_for('suppliers'))
    return render_template('edit_supplier.html', form=form, supplier=supplier)

# ---------- Purchase (Stock In) ----------
@app.route('/stock_in', methods=['GET','POST'])
@login_required
@roles_required('admin','pharmacist')
def stock_in():
    branch = get_active_branch()
    if not branch:
        return redirect(url_for('select_branch'))
    form = PurchaseForm()
    form.product_id.choices = [(p.id, f'{p.name} ({p.quantity} stock)') for p in Product.query.filter_by(branch_id=branch.id).order_by(Product.name).all()]
    form.supplier_id.choices = [(0, 'None')] + [(s.id, s.name) for s in Supplier.query.all()]
    if form.validate_on_submit():
        product = Product.query.get(form.product_id.data)
        if not product or product.branch_id != branch.id:
            flash('Invalid product.')
            return redirect(url_for('stock_in'))
        purchase = Purchase(
            branch_id=branch.id,
            supplier_id=form.supplier_id.data if form.supplier_id.data != 0 else None,
            product_id=product.id,
            quantity=form.quantity.data,
            buying_price=form.buying_price.data
        )
        product.quantity += form.quantity.data
        product.purchase_price = form.buying_price.data
        db.session.add(purchase)
        db.session.commit()
        flash('Stock added.')
        return redirect(url_for('products'))
    return render_template('stock_in.html', form=form, branch=branch)

# ---------- Stock Transfer ----------
@app.route('/stock_transfer', methods=['GET','POST'])
@login_required
@roles_required('admin')
def stock_transfer():
    form = StockTransferForm()
    branches = Branch.query.all()
    form.from_branch_id.choices = [(b.id, b.name) for b in branches]
    form.to_branch_id.choices = [(b.id, b.name) for b in branches]
    form.product_id.choices = [(p.id, f'{p.name} (from branch {p.branch.name})') for p in Product.query.all()]
    if form.validate_on_submit():
        if form.from_branch_id.data == form.to_branch_id.data:
            flash('Cannot transfer to same branch.')
            return render_template('stock_transfer.html', form=form)
        product = Product.query.get(form.product_id.data)
        if not product or product.branch_id != form.from_branch_id.data:
            flash('Product not found in origin branch.')
            return render_template('stock_transfer.html', form=form)
        if product.quantity < form.quantity.data:
            flash('Not enough stock.')
            return render_template('stock_transfer.html', form=form)
        product.quantity -= form.quantity.data
        dest_product = Product.query.filter_by(branch_id=form.to_branch_id.data, barcode=product.barcode).first()
        if not dest_product:
            dest_product = Product(
                branch_id=form.to_branch_id.data,
                name=product.name,
                barcode=product.barcode,
                batch_number=product.batch_number,
                category_id=product.category_id,
                supplier_id=product.supplier_id,
                purchase_price=product.purchase_price,
                selling_price=product.selling_price,
                quantity=0,
                expiry_date=product.expiry_date
            )
            db.session.add(dest_product)
        dest_product.quantity += form.quantity.data
        transfer = StockTransfer(
            from_branch_id=form.from_branch_id.data,
            to_branch_id=form.to_branch_id.data,
            product_id=product.id,
            quantity=form.quantity.data
        )
        db.session.add(transfer)
        db.session.commit()
        flash('Transfer successful.')
        return redirect(url_for('products'))
    return render_template('stock_transfer.html', form=form)

# ---------- Expiry Alerts ----------
@app.route('/expiry_alerts')
@login_required
def expiry_alerts():
    branch = get_active_branch()
    if not branch:
        return redirect(url_for('select_branch'))
    today = date.today()
    ninety_days = today + timedelta(days=90)
    expiring = Product.query.filter(
        Product.branch_id==branch.id,
        Product.expiry_date <= ninety_days,
        Product.expiry_date >= today
    ).order_by(Product.expiry_date).all()
    return render_template('expiry_alerts.html', branch=branch, products=expiring, now=today)

# ---------- Sales History ----------
@app.route('/sales_history')
@login_required
def sales_history():
    branch = get_active_branch()
    if not branch:
        return redirect(url_for('select_branch'))
    page = request.args.get('page', 1, type=int)
    start = request.args.get('start_date')
    end = request.args.get('end_date')
    query = Sale.query.filter_by(branch_id=branch.id).order_by(Sale.sale_date.desc())
    if start:
        query = query.filter(db.func.date(Sale.sale_date) >= start)
    if end:
        query = query.filter(db.func.date(Sale.sale_date) <= end)
    sales = query.paginate(page=page, per_page=20)
    return render_template('sales_history.html', branch=branch, sales=sales, start_date=start, end_date=end)

# ---------- Reports ----------
@app.route('/reports')
@login_required
@roles_required('admin')
def reports():
    branch = get_active_branch()
    if not branch:
        return redirect(url_for('select_branch'))
    today = date.today()
    daily_sales = db.session.query(db.func.sum(Sale.total_amount)).filter(Sale.branch_id==branch.id, db.func.date(Sale.sale_date)==today).scalar() or 0
    daily_profit = db.session.query(db.func.sum(SaleItem.profit)).join(Sale).filter(Sale.branch_id==branch.id, db.func.date(Sale.sale_date)==today).scalar() or 0
    return render_template('reports.html', branch=branch, daily_sales=daily_sales, daily_profit=daily_profit)

# ---------- User Management ----------
@app.route('/users')
@login_required
@roles_required('admin')
def users():
    users = User.query.all()
    return render_template('users.html', users=users)

@app.route('/add_user', methods=['GET','POST'])
@login_required
@roles_required('admin')
def add_user():
    form = UserForm()
    form.role_id.choices = [(r.id, r.name) for r in Role.query.all()]
    form.branch_id.choices = [(0, 'None')] + [(b.id, b.name) for b in Branch.query.all()]
    if form.validate_on_submit():
        user = User(
            username=form.username.data,
            password_hash=generate_password_hash(form.password.data),
            role_id=form.role_id.data,
            branch_id=form.branch_id.data if form.branch_id.data != 0 else None,
            is_active=form.is_active.data
        )
        # Generate TOTP secret for the new user
        user.totp_secret = pyotp.random_base32()
        db.session.add(user)
        db.session.commit()
        flash('User created. 2FA secret generated. Show QR code below.')
        return redirect(url_for('setup_2fa', user_id=user.id))
    return render_template('add_user.html', form=form)

@app.route('/toggle_user_status/<int:user_id>', methods=['POST'])
@login_required
@roles_required('admin')
def toggle_user_status(user_id):
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        return jsonify({'error': 'Cannot change your own status'}), 400
    data = request.get_json()
    user.is_active = data.get('is_active', False)
    try:
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Failed to update user status'}), 500

@app.route('/delete_user/<int:user_id>', methods=['DELETE'])
@login_required
@roles_required('admin')
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        return jsonify({'error': 'Cannot delete your own account'}), 400
    sales_count = Sale.query.filter_by(user_id=user_id).count()
    if sales_count > 0:
        return jsonify({'error': 'Cannot delete user with sales history.'}), 400
    try:
        db.session.delete(user)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Failed to delete user'}), 500

# ---------- Categories CRUD ----------
@app.route('/categories')
@login_required
@roles_required('admin','pharmacist')
def categories():
    categories = Category.query.order_by(Category.name).all()
    return render_template('categories.html', categories=categories)

@app.route('/add_category', methods=['GET','POST'])
@login_required
@roles_required('admin','pharmacist')
def add_category():
    form = CategoryForm()
    if form.validate_on_submit():
        category = Category(name=form.name.data, description=form.description.data)
        db.session.add(category)
        db.session.commit()
        flash('Category created.')
        return redirect(url_for('categories'))
    return render_template('add_category.html', form=form)

@app.route('/add_category_quick', methods=['POST'])
@login_required
@roles_required('admin','pharmacist')
def add_category_quick():
    data = request.get_json()
    if not data or not data.get('name'):
        return jsonify({'error': 'Category name is required'}), 400
    existing = Category.query.filter_by(name=data['name'].strip()).first()
    if existing:
        return jsonify({'error': 'Category already exists'}), 400
    try:
        category = Category(name=data['name'].strip(), description=data.get('description', '').strip())
        db.session.add(category)
        db.session.commit()
        return jsonify({'success': True, 'category_id': category.id, 'category_name': category.name})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Failed to create category'}), 500

@app.route('/edit_category/<int:id>', methods=['GET','POST'])
@login_required
@roles_required('admin','pharmacist')
def edit_category(id):
    category = Category.query.get_or_404(id)
    form = CategoryForm(obj=category)
    if form.validate_on_submit():
        category.name = form.name.data
        category.description = form.description.data
        db.session.commit()
        flash('Category updated.')
        return redirect(url_for('categories'))
    return render_template('edit_category.html', form=form, category=category)

@app.route('/delete_category/<int:category_id>', methods=['DELETE'])
@login_required
@roles_required('admin','pharmacist')
def delete_category(category_id):
    category = Category.query.get_or_404(category_id)
    if category.products:
        return jsonify({'error': 'Cannot delete category with products.'}), 400
    try:
        db.session.delete(category)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'Failed to delete category'}), 500

# ---------- 2FA Setup Routes ----------
@app.route('/setup_2fa/<int:user_id>')
@login_required
@roles_required('admin')
def setup_2fa(user_id):
    user = User.query.get_or_404(user_id)
    if not user.totp_secret:
        flash('No 2FA secret for this user.')
        return redirect(url_for('users'))
    totp = pyotp.TOTP(user.totp_secret)
    provisioning_uri = totp.provisioning_uri(name=user.username, issuer_name='SAMS Clinic POS')
    img = qrcode.make(provisioning_uri)
    buffer = io.BytesIO()
    img.save(buffer, format='PNG')
    img_str = base64.b64encode(buffer.getvalue()).decode()
    return render_template('setup_2fa.html', user=user, qr_code=img_str, secret=user.totp_secret)

@app.route('/enable_2fa/<int:user_id>')
@login_required
@roles_required('admin')
def enable_2fa(user_id):
    user = User.query.get_or_404(user_id)
    if not user.totp_secret:
        user.totp_secret = pyotp.random_base32()
        db.session.commit()
    return redirect(url_for('setup_2fa', user_id=user.id))

@app.route('/reset_2fa/<int:user_id>')
@login_required
@roles_required('admin')
def reset_2fa(user_id):
    user = User.query.get_or_404(user_id)
    user.totp_secret = pyotp.random_base32()
    db.session.commit()
    flash('2FA secret reset. Show the new QR code to the user.')
    return redirect(url_for('setup_2fa', user_id=user.id))

# ---------- Startup ----------
@app.before_request
def set_default_branch():
    if current_user.is_authenticated and not session.get('active_branch_id') and not current_user.is_admin():
        if current_user.branch:
            session['active_branch_id'] = current_user.branch.id

# ---------- Export Routes ----------
@app.route('/export/<report_type>/<format>')
@login_required
@roles_required('admin')
def export_report(report_type, format):
    branch = get_active_branch()
    if not branch:
        flash('Select a branch first.')
        return redirect(url_for('select_branch'))
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    if report_type == 'sales':
        return export_sales(branch, format, start_date, end_date)
    elif report_type == 'profit':
        return export_profit(branch, format, start_date, end_date)
    elif report_type == 'inventory':
        return export_inventory(branch, format)
    elif report_type == 'expiry':
        return export_expiry(branch, format)
    else:
        flash('Invalid report type.')
        return redirect(url_for('reports'))

def export_sales(branch, format, start_date=None, end_date=None):
    query = Sale.query.filter(Sale.branch_id == branch.id)
    if start_date:
        query = query.filter(Sale.sale_date >= datetime.strptime(start_date, '%Y-%m-%d'))
    if end_date:
        query = query.filter(Sale.sale_date <= datetime.strptime(end_date, '%Y-%m-%d'))
    sales = query.order_by(Sale.sale_date.desc()).all()
    if format == 'pdf':
        return create_sales_pdf(sales, branch, start_date, end_date)
    elif format == 'excel':
        return create_sales_excel(sales, branch, start_date, end_date)

def export_profit(branch, format, start_date=None, end_date=None):
    query = SaleItem.query.join(Sale).filter(Sale.branch_id == branch.id)
    if start_date:
        query = query.filter(Sale.sale_date >= datetime.strptime(start_date, '%Y-%m-%d'))
    if end_date:
        query = query.filter(Sale.sale_date <= datetime.strptime(end_date, '%Y-%m-%d'))
    sale_items = query.all()
    if format == 'pdf':
        return create_profit_pdf(sale_items, branch, start_date, end_date)
    elif format == 'excel':
        return create_profit_excel(sale_items, branch, start_date, end_date)

def export_inventory(branch, format):
    products = Product.query.filter_by(branch_id=branch.id).order_by(Product.name).all()
    if format == 'pdf':
        return create_inventory_pdf(products, branch)
    elif format == 'excel':
        return create_inventory_excel(products, branch)

def export_expiry(branch, format):
    today = date.today()
    expiring_products = Product.query.filter(
        Product.branch_id == branch.id,
        Product.expiry_date <= today + timedelta(days=90)
    ).order_by(Product.expiry_date).all()
    if format == 'pdf':
        return create_expiry_pdf(expiring_products, branch)
    elif format == 'excel':
        return create_expiry_excel(expiring_products, branch)

# ---------- PDF / Excel Creation Functions ----------
def create_sales_pdf(sales, branch, start_date, end_date):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []
    
    title = Paragraph(f"SAMS Clinic - Sales Report<br/>{branch.name}<br/>Date Range: {start_date or 'All'} to {end_date or 'All'}", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    data = [['Receipt #', 'Date', 'Cashier', 'Total', 'Payment Method']]
    for sale in sales:
        data.append([
            sale.receipt_number,
            sale.sale_date.strftime('%Y-%m-%d %H:%M'),
            sale.user.username,
            f"KES {sale.total_amount:.2f}",
            sale.payment_method
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=sales_report_{branch.name}_{date.today()}.pdf'
    return response

def create_sales_excel(sales, branch, start_date, end_date):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    headers = ['Receipt #', 'Date', 'Cashier', 'Total', 'Payment Method']
    ws.append(headers)
    
    for sale in sales:
        ws.append([
            sale.receipt_number,
            sale.sale_date.strftime('%Y-%m-%d %H:%M'),
            sale.user.username,
            sale.total_amount,
            sale.payment_method
        ])
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    response = make_response(openpyxl.writer.excel.save_virtual_workbook(wb))
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=sales_report_{branch.name}_{date.today()}.xlsx'
    return response

def create_profit_pdf(sale_items, branch, start_date, end_date):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []
    
    title = Paragraph(f"SAMS Clinic - Profit Report<br/>{branch.name}<br/>Date Range: {start_date or 'All'} to {end_date or 'All'}", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    data = [['Product', 'Quantity', 'Unit Price', 'Unit Cost', 'Profit']]
    total_profit = 0
    for item in sale_items:
        profit = item.profit
        total_profit += profit
        data.append([
            item.product.name,
            item.quantity,
            f"KES {item.price:.2f}",
            f"KES {item.price - (profit/item.quantity):.2f}",
            f"KES {profit:.2f}"
        ])
    
    data.append(['', '', '', 'TOTAL PROFIT:', f"KES {total_profit:.2f}"])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-2, -1), colors.beige),
        ('BACKGROUND', (-1, -1), (-1, -1), colors.lightgreen),
        ('FONTNAME', (-1, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=profit_report_{branch.name}_{date.today()}.pdf'
    return response

def create_profit_excel(sale_items, branch, start_date, end_date):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Profit Report"
    
    headers = ['Product', 'Quantity', 'Unit Price', 'Unit Cost', 'Profit']
    ws.append(headers)
    
    total_profit = 0
    for item in sale_items:
        profit = item.profit
        total_profit += profit
        ws.append([
            item.product.name,
            item.quantity,
            item.price,
            item.price - (profit/item.quantity),
            profit
        ])
    
    ws.append(['', '', '', 'TOTAL PROFIT:', total_profit])
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    total_row = len(sale_items) + 2
    for col in range(4, 6):
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    
    response = make_response(openpyxl.writer.excel.save_virtual_workbook(wb))
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=profit_report_{branch.name}_{date.today()}.xlsx'
    return response

def create_inventory_pdf(products, branch):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []
    
    title = Paragraph(f"SAMS Clinic - Inventory Report<br/>{branch.name}<br/>Generated: {date.today()}", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    data = [['Product', 'Category', 'Barcode', 'Batch', 'Quantity', 'Expiry Date']]
    for product in products:
        data.append([
            product.name,
            product.category.name if product.category else 'N/A',
            product.barcode or 'N/A',
            product.batch_number or 'N/A',
            product.quantity,
            product.expiry_date.strftime('%Y-%m-%d')
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=inventory_report_{branch.name}_{date.today()}.pdf'
    return response

def create_inventory_excel(products, branch):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Report"
    
    headers = ['Product', 'Category', 'Barcode', 'Batch', 'Quantity', 'Expiry Date']
    ws.append(headers)
    
    for product in products:
        ws.append([
            product.name,
            product.category.name if product.category else 'N/A',
            product.barcode or 'N/A',
            product.batch_number or 'N/A',
            product.quantity,
            product.expiry_date.strftime('%Y-%m-%d')
        ])
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    response = make_response(openpyxl.writer.excel.save_virtual_workbook(wb))
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=inventory_report_{branch.name}_{date.today()}.xlsx'
    return response

def create_expiry_pdf(products, branch):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []
    
    title = Paragraph(f"SAMS Clinic - Expiry Report<br/>{branch.name}<br/>Generated: {date.today()}", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    data = [['Product', 'Category', 'Quantity', 'Expiry Date', 'Days Until Expiry']]
    today = date.today()
    for product in products:
        days_until = (product.expiry_date - today).days
        data.append([
            product.name,
            product.category.name if product.category else 'N/A',
            product.quantity,
            product.expiry_date.strftime('%Y-%m-%d'),
            f"{days_until} days"
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=expiry_report_{branch.name}_{date.today()}.pdf'
    return response

def create_expiry_excel(products, branch):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expiry Report"
    
    headers = ['Product', 'Category', 'Quantity', 'Expiry Date', 'Days Until Expiry']
    ws.append(headers)
    
    today = date.today()
    for product in products:
        days_until = (product.expiry_date - today).days
        ws.append([
            product.name,
            product.category.name if product.category else 'N/A',
            product.quantity,
            product.expiry_date.strftime('%Y-%m-%d'),
            days_until
        ])
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for row in range(2, len(products) + 2):
        days_cell = ws.cell(row=row, column=5)
        days = days_cell.value
        if days < 0:
            fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        elif days < 30:
            fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        else:
            fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = fill
    
    response = make_response(openpyxl.writer.excel.save_virtual_workbook(wb))
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=expiry_report_{branch.name}_{date.today()}.xlsx'
    return response

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        if not Role.query.first():
            admin_role = Role(name='admin')
            pharmacist_role = Role(name='pharmacist')
            cashier_role = Role(name='cashier')
            db.session.add_all([admin_role, pharmacist_role, cashier_role])
            db.session.commit()
            branch1 = Branch(name='SAMS Clinic A', location='Main Street')
            branch2 = Branch(name='SAMS Clinic B', location='Market Road')
            branch3 = Branch(name='SAMS Clinic C', location='Hospital Lane')
            db.session.add_all([branch1, branch2, branch3])
            db.session.commit()
            admin = User(username='admin', password_hash=generate_password_hash('admin123'), role_id=admin_role.id, is_active=True)
            pharmacist = User(username='pharm', password_hash=generate_password_hash('pharm123'), role_id=pharmacist_role.id, branch_id=branch1.id, is_active=True)
            cashier = User(username='cash', password_hash=generate_password_hash('cash123'), role_id=cashier_role.id, branch_id=branch1.id, is_active=True)
            db.session.add_all([admin, pharmacist, cashier])
            db.session.commit()
            categories = [
                Category(name='Medications', description='Prescription and over-the-counter medications'),
                Category(name='Medical Supplies', description='Medical equipment and supplies'),
                Category(name='Personal Care', description='Personal hygiene and care products'),
                Category(name='Vitamins & Supplements', description='Dietary supplements and vitamins'),
                Category(name='First Aid', description='First aid supplies and emergency care'),
                Category(name='Baby Care', description='Baby products and infant care'),
                Category(name='Health Devices', description='Medical monitoring devices'),
                Category(name='Herbal Products', description='Traditional and herbal medicines')
            ]
            db.session.add_all(categories)
            db.session.commit()
    app.run(debug=True)